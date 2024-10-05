import os
import sys
import logging
from logging.handlers import RotatingFileHandler
import requests
import openpyxl
import json
from datetime import datetime, date
from dateutil.relativedelta import relativedelta
from utils import base_url, months_br
import time

# Processa as configurações do helix .config
def load_configurations(config_file):
    """
    Carrega configurações de helix_id e api_key de um arquivo JSON.

    Args:
        config_file (str): Caminho para o arquivo de configuração JSON.

    Returns:
        list: Lista de dicionários contendo helix_id e api_key.
    """
    try:
        with open(config_file, 'r') as file:
            configs = json.load(file)
        return configs.get('helix_configs', [])
    except FileNotFoundError:
        logging.warning(f"Arquivo de configuração não encontrado: {config_file}")
    except json.JSONDecodeError as e:
        logging.warning(f"Erro ao decodificar JSON no arquivo de configuração: {e}")
    except Exception as e:
        logging.warning(f"Erro ao carregar configurações: {e}")
    return []

# Funçao que verifica se o arquivo existe na pasta
def file_exists(file_path):
    """
    Verifica se o arquivo já existe no caminho especificado.

    Args:
        file_path (str): Caminho completo para o arquivo a ser verificado.

    Returns:
        bool: True se o arquivo existe, False caso contrário.
    """
    return os.path.exists(file_path)

# Processa a configuração e chama a funçao para cada par de helix_id e api_key
def process_helix_configs(configs):
    """
    Processa cada configuração de helix_id e api_key.

    Args:
        configs (list): Lista de dicionários com helix_id e api_key.
    """
    for config in configs:
        company = config.get('company')
        helix_id = config.get('helix_id')
        api_key = config.get('api_key')
        if helix_id and api_key:
            process_helix(company, helix_id, api_key)
        else:
            logging.warning(f"{company} - Configuração inválida encontrada")

# Funçoes para realizar as buscas
def process_helix(company, helix_id, api_key):
    """
    Realiza operações para um par específico de helix_id e api_key.

    Args:
        helix_id (str): ID do Helix.
        api_key (str): Chave da API.
    """
    # Pega todas as pesquisas salvas
    def get_search_saved():
        """
        Busca todas as pesquisas salvas na API para um helix_id específico.
        """
        try:
            url = f"{base_url}/helix/id/{helix_id}/api/v3/search/saved/?limit=50"
            headers = {
                "x-fireeye-api-key": api_key,
                "accept": "application/json"
            }
            response = requests.get(url, headers=headers)
            response.raise_for_status()
            return response.json()
        except requests.exceptions.RequestException as e:
            logging.warning(f"{company} - Erro ao fazer a requisição para searches salvas: {e}")
        except Exception as e:
            logging.warning(f"{company} - Erro ao buscar searches salvas: {e}")
        return None
    
    # Processa a pesquisa por tipo
    def run_search(search):
        """
        Valida e executa a pesquisa com base no tipo (normal ou alerta).
        
        Args:
            search (dict): Dicionário contendo detalhes da pesquisa.
        """
        query = search.get('query')
        name = search.get('name')

        month_name = months_br.get(date.today().month, 'Desconhecido')

        save_path = os.path.join('dados', str(date.today().year), company, month_name)
        os.makedirs(save_path, exist_ok=True)

        filename = os.path.join(save_path, f"{name}.xlsx")

        if not os.path.exists(filename):
            try:
                if 'alerts' in query or 'WAF' in name or 'analytics' in query:
                    logging.info(f"{company} - Pesquisa search: {name}")
                    get_index(company, name, query)
                elif '04 - Origen User Top One Failure Login' in name: 
                    process_archive(company, name, failure_Login04(company, name))
                elif '05 - Origen User Top One Failure Login' in name: 
                    process_archive(company, name, failure_Login05(company, name))
                else:
                    process_archive(company,name, query)
            except Exception as e:
                logging.warning(f"{company} - Erro ao executar a pesquisa '{name}': {e}")
        else:
            logging.debug(f"Já existe o arquivo '{filename}' na pasta '{os.path.dirname(filename)}' com a query: {query}")
            
    # Processando index
    def get_index(company,name, query):
        """
        Executa a pesquisa no índice e salva os resultados.
        
        Args:
            name (str): Nome da pesquisa.
            query (str): Query da pesquisa.
        """
        try:
            url = f"https://apps.fireeye.com/helix/id/{helix_id}/api/v1/search"
            headers = {
                "Content-Type": "application/json",
                "x-fireeye-api-key": api_key
            }
            payload = {"query": f"start='1 month ago' {query}"}
            
            response = requests.post(url, json=payload, headers=headers)
            response.raise_for_status()

            results = response.json().get('results', {}).get('aggregations', {}).items()
            
            data = None
            for _, v in results:
                if 'buckets' in v:
                    data = v['buckets']
                    break

            if data:
                save_index(company, name, query, data)
            else:
                logging.warning(f"{company} - Sem dados para a pesquisa '{name}'")
        except requests.exceptions.RequestException as e:
            logging.warning(f"{company} - Erro na requisição ao endpoint '{name}': {e}")
        except Exception as e:
            logging.warning(f"{company} - Erro ao obter índice para a pesquisa '{name}': {e}")

    # Cria as tabelas
    def save_index(company, name, query, results):
        """
        Salva os resultados da pesquisa em uma planilha Excel no diretório 'dados/ano/empresa/mês'.
        
        Args:
            name (str): Nome da pesquisa.
            query (str): Query da pesquisa.
            results (list): Lista de resultados para salvar.
        """
        query_columns = query.split('groupby')[1].replace('[', '').replace(']', '').strip().split(',')
        try:
            
            month_name = months_br.get(date.today().month, 'Desconhecido')
            save_path = os.path.join('dados', str(date.today().year), company, month_name)
            os.makedirs(save_path, exist_ok=True)
            filename = os.path.join(save_path, f"{name}.xlsx")

            if not file_exists(filename):
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = name

                columns = query_columns + ['count']
                ws.append(columns)

                for result in results:
                    raw_key = result.get('key', '')
                    doc_count = result.get('doc_count', '')

                    processed_key = [segment.strip() for segment in raw_key.split('|%$,$%|')]

                    if len(processed_key) == len(columns) - 1:
                        processed_key.append(doc_count)
                        ws.append(processed_key)
                    else:
                        logging.warning(f"{company} - Formato inesperado na chave '{raw_key}': número de segmentos incorreto")

                wb.save(filename)
                logging.info(f"{company} - Pesquisa '{name}' salvo com sucesso em '{filename}'")
            else:
                logging.warning(f"{company} - Arquivo '{filename}' já existe.")
        except Exception as e:
            logging.warning(f"{company} - Erro ao salvar resultado da pesquisa '{name}': {e}")

    # Processando o archive
    def process_archive(company,name, query):
        """
        Executa os processo no archive, criar e buscar o resutlado.
        
        Args:
            company (str): Nome da empresa.
            name (str): Nome da pesquisa.
            query (str): Query da pesquisa.
        """
        # cria uma pesquisa no archive
        def create_archive(query):
            reqUrl = f"https://apps.fireeye.com/helix/id/{helix_id}/api/v1/search/archive/"
        
            headersList = {
                "Content-Type": "application/json",
                "x-fireeye-api-key": api_key
            }
        
            payload = get_date_range(query)
        
            response = requests.request(
                "POST", reqUrl, json=payload, headers=headersList)
        
            if response.status_code == 201:
                id = response.json()['data'][0]['id']
                logging.info(f"{company} - Pesquisa archive {name} - criada com ID: {id}")
                return id
            else:
                logging.warning(f"{company} - Erro ao criar pesquisa archive {name}: {response.text}")
                return None
        
        # Verifica o status da pesquisa pelo id
        def check_archive(id):

            reqUrl = f"https://apps.fireeye.com/helix/id/{helix_id}/api/v1/search/archive/{id}"

            headersList = {
                "Content-Type": "application/json",
                "x-fireeye-api-key": api_key
            }

            response = requests.request(
                "GET", reqUrl, headers=headersList)

            state = response.json()['data'][0]['state']

            return state

        # Busca o resultado da pesquisa pelo id
        def get_archive(id):
            try:
                reqUrl = f"https://apps.fireeye.com/helix/id/{helix_id}/api/v1/search/archive/{id}/results"

                headersList = {
                    "Content-Type": "application/json",
                    "x-fireeye-api-key": api_key
                }

                response = requests.request("GET", reqUrl, headers=headersList)
                response.raise_for_status()

                results = response.json()['results']['results']['aggregations'].items()
                
                data = None
                for _, v in results:
                    if 'buckets' in v:
                        data = v['buckets']
                        break

                if data:
                    save_index(company, name, query, data)
                    return data
                else:
                    logging.warning(f"{company} - Sem dados para a pesquisa '{name}'")
            except requests.exceptions.RequestException as e:
                logging.warning(f"{company} - Erro na requisição ao endpoint '{name}': {e}")
            except Exception as e:
                logging.warning(f"{company} - Erro ao obter índice para a pesquisa '{name}': {e}")

        # Executando o archive
        id = create_archive(query)
        if id:
            state = check_archive(id)
            while state != 'completed':
                time.sleep(60)
                state = check_archive(id)
            if state == 'completed':
                get_archive(id)
           
            
    # Main process for a single helix_id and api_key
    researches = get_search_saved()
    if researches:
        for search in researches.get('results', []):
            run_search(search)

# Funçao para deletar todas pesquisas
def delete_all(configs):
    """
    Deleta todas as pesquisas para cada configuração fornecida.

    Args:
        configs (list): Lista de dicionários com helix_id, api_key e company.
    """
    # Pega todas as pequisas para deletar
    def delete_all_archives(company, helix_id, api_key):
        """
        Deleta todas as pesquisas arquivadas para um helix_id específico.

        Args:
            company (str): Nome da empresa.
            helix_id (str): ID do Helix.
            api_key (str): Chave da API.

        Returns:
            bool: True se as pesquisas foram deletadas, False caso contrário.
        """
        resource = f"/helix/id/{helix_id}/api/v1/search/archive/"
        reqUrl = base_url + resource

        headers = {
            "Content-Type": "application/json",
            "x-fireeye-api-key": api_key
        }

        params = {"limit": "500"}

        try:
            response = requests.get(reqUrl, headers=headers, params=params)
            response.raise_for_status()

            data = response.json().get('data', [])
            count = response.json().get('meta', {}).get('totalCount', 0)

            if count > 0:
                logging.info(f"{company} - Iniciando exclusão de {count} pesquisas arquivadas.")
                for archive in data:
                    delete_archive(company, helix_id, api_key, archive['id'])
                return True
            else:
                logging.info(f"{company} - Nenhuma pesquisa arquivada encontrada para exclusão.")
                return False

        except requests.exceptions.RequestException as e:
            logging.warning(f"{company} - Erro na requisição para buscar arquivos arquivados: {e}")
        except Exception as e:
            logging.warning(f"{company} - Erro ao processar exclusão de arquivos arquivados: {e}")
        return False
    
    # Deleta uma pesquisa por id
    def delete_archive(company, helix_id, api_key, archive_id):
        """
        Deleta uma pesquisa arquivada específica.

        Args:
            company (str): Nome da empresa.
            helix_id (str): ID do Helix.
            api_key (str): Chave da API.
            archive_id (str): ID do arquivo a ser deletado.
        """
        resource = f'/helix/id/{helix_id}/api/v1/search/archive/{archive_id}'
        reqUrl = base_url + resource

        headers = {
            "Content-Type": "application/json",
            "x-fireeye-api-key": api_key
        }

        try:
            response = requests.delete(reqUrl, headers=headers)
            if response.status_code == 204:
                logging.info(f"{company} - Pesquisa {archive_id} deletada com sucesso.")
            else:
                logging.warning(f"{company} - Erro ao deletar pesquisa {archive_id}: {response.status_code}")
        except requests.exceptions.RequestException as e:
            logging.warning(f"{company} - Erro na requisição ao deletar pesquisa {archive_id}: {e}")
        except Exception as e:
            logging.warning(f"{company} - Erro ao deletar pesquisa {archive_id}: {e}")

    # loop para deletar todas as pesquisas
    for config in configs:
        company = config.get('company')
        helix_id = config.get('helix_id')
        api_key = config.get('api_key')

        if helix_id and api_key:
            if delete_all_archives(company, helix_id, api_key):
                logging.info(f"{company} - Todas as pesquisas arquivadas foram deletadas.")
        else:
            logging.warning(f"{company} - Configuração inválida encontrada: helix_id ou api_key ausentes.")

    return None

# Configuração de logs
def setup_logging():
    """
    Configura o logging para o script.
    """
    log_file = 'log_dev.log'
    formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
    handler = RotatingFileHandler(log_file, maxBytes=1000000, backupCount=5)
    handler.setFormatter(formatter)

    logger = logging.getLogger()
    logger.setLevel(logging.INFO)
    logger.addHandler(handler)

# EXTRAS
# verifica a data a ser pequisada
def get_date_range(query):
    """
    Seta as datas para o payload
    """
    c_day = date.today().day
    c_month = date.today().month
    c_year = date.today().year
    c_hour = datetime.now().hour
    
    if c_day > 1 and c_day < 15:
        search_start_date = f'{c_year}-{c_month-1}-01T00:00:00Z'
        search_end_date = f'{c_year}-{c_month-1}-30T23:59:59Z'
    else:
        search_start_date = f'{c_year}-{c_month}-01T00:00:00Z'
        search_end_date = f'{c_year}-{c_month}-{c_day}T{c_hour-4}:59:59Z'  


    payload = {
        "query": f"{query}",
        "searchEndDate": search_end_date,
        "searchStartDate": search_start_date
    }

    # print(payload)
    return payload


def failure_Login04(company, name):

    month_name = months_br.get(date.today().month, 'Desconhecido')

    save_path = os.path.join('dados', str(date.today().year), company, month_name)
    os.makedirs(save_path, exist_ok=True)

    filename = os.path.join(save_path, "03 - Failure Login.xlsx")

    pt = openpyxl.load_workbook(filename)
    pl = pt.active
    user = pl['A2'].value
    user = user.lstrip("0")

    query = f'metaclass=windows has=srcipv4 NOT targetusername:["*$"] targetusername:[`{user}`] eventid=4625 | groupby srcipv4'
    
    return query


def failure_Login05(company, name):

    month_name = months_br.get(date.today().month, 'Desconhecido')

    save_path = os.path.join('dados', str(date.today().year), company, month_name)
    os.makedirs(save_path, exist_ok=True)

    filename = os.path.join(save_path, "03 - Failure Login.xlsx")

    pt = openpyxl.load_workbook(filename)
    pl = pt.active
    user = pl['A2'].value
    user = user.lstrip("0")

    query = f'metaclass=windows targetusername:[`{user}`] eventid=4625 | groupby meta_sip4'
    
    return query
# END EXTRAS

# Display do menu
def display_menu():
    """
    Exibe o menu de opções para o usuário.
    """
    print("\n----- Menu -----")
    print("1. Processar Pesquisas")
    print("2. Escluir Todas as Pesquisas")
    print("0. Sair")
    choice = input("Escolha uma opção: ")
    return choice

# Função principal
def main():
    """
    Função principal que carrega configurações e processa cada par de helix_id e api_key.
    """
    setup_logging()
    config_file = 'config.json'
    
    while True:
        os.system('clear')
        choice = display_menu()
        if choice == '1':
            configs = load_configurations(config_file)
            process_helix_configs(configs)
        elif choice == '2':
            configs = load_configurations(config_file)
            delete_all(configs)
        elif choice == '0':
            break
        else:
            print("Opcão inválida. Tente novamente.")
        
if __name__ == "__main__":
    main()
